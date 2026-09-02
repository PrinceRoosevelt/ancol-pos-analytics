import gzip
import math
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path
from threading import Lock
from typing import Any

# Fix Windows console encoding for emoji characters
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

from database import (
    fetch_all_sales,
    fetch_all_targets,
    fetch_all_visitors,
    save_visitor_db,
    sync_database,
)


BASE_DIR = Path(__file__).resolve().parent
# Sales source root using the NEW POS export format.
# The parser automatically discovers every year folder named:
#   data/sales detail 2025
#   data/sales detail 2026
#   data/sales detail 2027
# etc.
# Add new monthly Excel files into the appropriate year folder and they
# will be picked up automatically. No code change is required per month.
SALES_ROOT_DIR = BASE_DIR / "data"
SALES_DETAIL_DIR_PATTERN = re.compile(r"^sales\s+detail\s+(\d{4})$", re.I)
REPORT_DIR = BASE_DIR / "reports"
MAPPING_FILE = BASE_DIR / "config" / "MASTER_OUTLET_MAPPING_V2.xlsx"
BUDGET_FILE = BASE_DIR / "data" / "target 2026" / "BUDGET_MERCH_ONLY_PYTHON_READY.xlsx"
# New POS export format:
#   A = No. Invoice (invoice header) / Kode Barang (item)
#   B = Nama Barang
#   C = Tgl & Jam Invoice (date on invoice header)
#   E = time on invoice header
#   G = Jml
#   I = Penjualan / BKP (DPP) at item level
#   N = Disc. Inv. at "Total Struk" level only
#   O = PPN
#   R = Total at "Total Struk" level
#
# IMPORTANT:
# - Column I is preserved as the original item-level Net Sales source.
# - "Total Struk" is NEVER treated as an item.
# - Transaction Net Sales is calculated from column R / 1.11.
# - The transaction-level Net Sales is allocated proportionally back to
#   item rows so product/outlet totals remain accurate after invoice discount.
SALES_VALUE_COLUMN_INDEX = 8       # I
SALES_VALUE_COLUMN_HEADER = "Penjualan"
TRANSACTION_TOTAL_COLUMN_INDEX = 17  # R
TRANSACTION_TOTAL_TAX_FACTOR = 1.11
INVOICE_DISCOUNT_COLUMN_INDEX = 13   # N, informational only
app = Flask(__name__)
_sales_cache: list[dict[str, Any]] | None = None
_sales_cache_signature: tuple[tuple[str, int, int], ...] = ()
_sales_cache_mapping_signature: tuple[int, int] | None = None
_sales_cache_lock = Lock()
_budget_cache: list[dict[str, Any]] | None = None
_budget_cache_signature: tuple[tuple[int, int], tuple[int, int]] | None = None
_budget_cache_lock = Lock()
MONTH_NAMES = {
    "01": "Januari",
    "02": "Februari",
    "03": "Maret",
    "04": "April",
    "05": "Mei",
    "06": "Juni",
    "07": "Juli",
    "08": "Agustus",
    "09": "September",
    "10": "Oktober",
    "11": "November",
    "12": "Desember",
}
WEEKDAY_NAMES = {
    0: "Senin",
    1: "Selasa",
    2: "Rabu",
    3: "Kamis",
    4: "Jumat",
    5: "Sabtu",
    6: "Minggu",
}


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value) if math.isfinite(float(value)) else 0.0
    cleaned = re.sub(r"[^\d,.-]", "", str(value).strip())
    if not cleaned:
        return 0.0
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    for fmt in ("%d %b %Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return None


def read_outlet_mapping() -> dict[str, str]:
    sheet = load_workbook(MAPPING_FILE, read_only=True, data_only=True)["OUTLET_MAPPING"]
    mapping: dict[str, str] = {}
    try:
        for outlet, area in sheet.iter_rows(min_row=2, values_only=True):
            if outlet and area:
                mapping[str(outlet).strip()] = str(area).strip()
    finally:
        sheet.parent.close()
    return mapping


def read_budget_daily() -> list[dict[str, Any]]:
    """Membaca daily targets langsung dari database SQLite (sudah sinkron dengan Excel)."""
    return fetch_all_targets()


def read_target_daily(
    month: str | None = None,
    date: str | None = None,
    outlet: str | None = None,
    area: str | None = None,
    available_dates: set[str] | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
) -> dict[str, float]:
    daily: dict[str, float] = defaultdict(float)
    start_dm = start_date[5:] if start_date else ""
    end_dm = end_date[5:] if end_date else ""
    for target_row in read_budget_daily():
        target_date = target_row["date"]
        target_dm = target_date[5:]
        if start_date or end_date:
            if start_dm and target_dm < start_dm:
                continue
            if end_dm and target_dm > end_dm:
                continue
        elif month and target_row["month"] != month:
            continue
        if date and target_date != date:
            continue
        if available_dates is not None and target_date not in available_dates:
            continue
        if outlet and target_row["outlet"] != outlet:
            continue
        if area and target_row["area"] != area:
            continue
        daily[target_date] += target_row["target"]
    return dict(daily)


def read_all_targets() -> list[dict[str, Any]]:
    return fetch_all_targets()


def read_all_visitors() -> list[dict[str, Any]]:
    """Baca data master pengunjung harian langsung dari database SQLite."""
    return fetch_all_visitors()


def save_visitor_actual(entry_date: str, unit: str, count: int) -> bool:
    """Simpan/Update nilai pengunjung harian pada sheet Excel dan database SQLite."""
    if not BUDGET_FILE.exists():
        return False
    dt_val = _date(entry_date)
    if not dt_val:
        raise ValueError(f"Format tanggal tidak valid: {entry_date}")

    # 1. Update ke file Excel
    wb = load_workbook(BUDGET_FILE, data_only=False)
    try:
        if "VISITOR_ACTUAL" not in wb.sheetnames:
            ws = wb.create_sheet("VISITOR_ACTUAL")
            ws.append(["Date", "Visitor Unit", "Visitor Actual"])
        else:
            ws = wb["VISITOR_ACTUAL"]

        target_date_str = dt_val.strftime("%Y-%m-%d")
        updated = False

        for r in range(2, ws.max_row + 1):
            cell_d = ws.cell(r, 1).value
            cell_u = ws.cell(r, 2).value
            if cell_d is not None and cell_u is not None:
                parsed_cell_d = _date(cell_d)
                if parsed_cell_d and parsed_cell_d.strftime("%Y-%m-%d") == target_date_str:
                    if str(cell_u).strip().casefold() == unit.strip().casefold():
                        ws.cell(r, 3, int(count))
                        updated = True
                        break

        if not updated:
            ws.append([dt_val, unit.strip(), int(count)])

        wb.save(BUDGET_FILE)
    finally:
        wb.close()

    # 2. Update ke Database SQLite
    area_name = VISITOR_UNIT_TO_AREA.get(unit.upper(), "")
    save_visitor_db(target_date_str, unit.strip(), int(count), area_name)

    _invalidate_sales_cache()
    return True


def read_sales() -> list[dict[str, Any]]:
    """Baca data sales langsung dari SQLite Database yang terindeks dan cepat."""
    global _sales_cache
    with _sales_cache_lock:
        if _sales_cache is not None:
            return _sales_cache

        # Pastikan file Excel yang baru / diubah tersinkronisasi ke DB
        sync_database()
        rows = fetch_all_sales()

        _sales_cache = rows
        return rows


def summarize(rows: list[dict[str, Any]], year: int) -> dict[str, Any]:
    selected = [row for row in rows if row["year"] == year]
    qty = sum(row["qty"] for row in selected)
    sales = sum(row["net_sales"] for row in selected)
    item_net_sales = sum(row.get("item_net_sales", 0.0) for row in selected)
    unique_tx = len({(r["outlet"], r["date"], r.get("invoice", "")) for r in selected})
    transactions = unique_tx if unique_tx > 0 else len(selected)
    atv = (sales / transactions) if transactions > 0 else 0.0
    upt = (qty / transactions) if transactions > 0 else 0.0
    asp = (sales / qty) if qty > 0 else 0.0
    return {
        "qty": qty,
        "net_sales": sales,
        "item_net_sales": item_net_sales,
        "transactions": transactions,
        "atv": atv,
        "upt": upt,
        "asp": asp,
        "items_count": len(selected),
    }


def growth_percent(current: float, previous: float) -> float | None:
    if previous == 0:
        return None
    return ((current - previous) / previous) * 100


def build_dashboard(
    rows: list[dict[str, Any]],
    month: str | None = None,
    date: str | None = None,
    outlet: str | None = None,
    area: str | None = None,
    include_raw: bool = True,
    start_date: str | None = None,
    end_date: str | None = None,
) -> dict[str, Any]:
    # The 2025 comparison must cover exactly the calendar days already
    # available in 2026. This prevents, for example, comparing 1-29 Aug 2026
    # against a full 1-31 Aug 2025 period.
    start_dm = start_date[5:] if start_date else ""
    end_dm = end_date[5:] if end_date else ""

    current_period_rows = [
        row
        for row in rows
        if row["year"] == 2026
        and (not start_date or row["date"] >= start_date)
        and (not end_date or row["date"] <= end_date)
        and (start_date or end_date or not month or row["month"] == month)
        and (not date or row["date"] == date)
    ]
    current_period_dates = {row["date"] for row in current_period_rows}
    all_sales_dates_2026 = {row["date"] for row in rows if row["year"] == 2026}
    all_sales_day_months_2026 = {target_date[5:] for target_date in all_sales_dates_2026}
    comparison_day_months = {target_date[5:] for target_date in current_period_dates}

    filtered = [
        row
        for row in rows
        if (
            (
                (start_date or end_date)
                and (not start_dm or row["date"][5:] >= start_dm)
                and (not end_dm or row["date"][5:] <= end_dm)
            )
            or (
                not (start_date or end_date)
                and (not month or row["month"][5:] == month[5:])
            )
        )
        and (not date or row["date"][5:] == date[5:])
        and (not outlet or row["outlet"] == outlet)
        and (not area or row["area"] == area)
        and (row["year"] == 2026 or row["date"][5:] in comparison_day_months)
    ]
    summary = {str(year): summarize(filtered, year) for year in (2025, 2026)}
    summary["growth"] = {
        key: growth_percent(summary["2026"][key], summary["2025"][key])
        for key in ("net_sales", "qty", "transactions", "atv", "upt", "asp")
    }
    summary["diff"] = {
        "net_sales": summary["2026"]["net_sales"] - summary["2025"]["net_sales"],
        "transactions": summary["2026"]["transactions"] - summary["2025"]["transactions"],
        "qty": summary["2026"]["qty"] - summary["2025"]["qty"],
        "atv": summary["2026"]["atv"] - summary["2025"]["atv"],
    }

    by_outlet: dict[str, dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: {"qty": 0.0, "net_sales": 0.0, "transactions": 0})
    )
    by_product: dict[str, dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: {"qty": 0.0, "net_sales": 0.0, "transactions": 0})
    )
    by_month: dict[str, dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: {"qty": 0.0, "net_sales": 0.0, "transactions": 0})
    )
    by_day: dict[str, dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: {"qty": 0.0, "net_sales": 0.0, "transactions": 0})
    )
    by_hour: dict[str, dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: {"qty": 0.0, "net_sales": 0.0, "transactions": 0})
    )
    by_area: dict[str, dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: {"qty": 0.0, "net_sales": 0.0, "transactions": 0})
    )
    outlet_area_map: dict[str, str] = {}

    for row in filtered:
        year_key = str(row["year"])
        day_key = row["date"][5:]
        outlet_area_map[row["outlet"]] = row["area"]

        for target_group, key in (
            (by_outlet, row["outlet"]),
            (by_product, row["product"]),
            (by_month, row["month"]),
            (by_day, day_key),
            (by_hour, row["hour"]),
            (by_area, row["area"]),
        ):
            target_group[key][year_key]["qty"] += row["qty"]
            target_group[key][year_key]["net_sales"] += row["net_sales"]
            if row["invoice"]:
                target_group[key][year_key][f"_inv_{row['invoice']}"] = 1

    for target_group in (by_outlet, by_product, by_month, by_day, by_hour, by_area):
        for key in target_group:
            for year_key in ("2025", "2026"):
                inv_keys = [k for k in target_group[key][year_key] if k.startswith("_inv_")]
                target_group[key][year_key]["transactions"] = len(inv_keys)
                for k in inv_keys:
                    del target_group[key][year_key][k]

    def flatten(target_group: dict[str, dict[str, dict[str, float]]]) -> list[dict[str, Any]]:
        result = []
        for name, values in sorted(target_group.items()):
            row_2026 = values.get("2026", {"qty": 0.0, "net_sales": 0.0, "transactions": 0})
            row_2025 = values.get("2025", {"qty": 0.0, "net_sales": 0.0, "transactions": 0})
            sales_2026 = row_2026["net_sales"]
            sales_2025 = row_2025["net_sales"]
            qty_2026 = row_2026["qty"]
            qty_2025 = row_2025["qty"]
            tx_2026 = row_2026["transactions"]
            tx_2025 = row_2025["transactions"]

            result.append(
                {
                    "name": name,
                    "2026": row_2026,
                    "2025": row_2025,
                    "growth_sales": growth_percent(sales_2026, sales_2025),
                    "growth_qty": growth_percent(qty_2026, qty_2025),
                    "growth_tx": growth_percent(tx_2026, tx_2025),
                    "diff_sales": sales_2026 - sales_2025,
                    "diff_qty": qty_2026 - qty_2025,
                    "diff_tx": tx_2026 - tx_2025,
                    "contrib_2026": (
                        (sales_2026 / summary["2026"]["net_sales"] * 100)
                        if summary["2026"]["net_sales"] > 0
                        else 0.0
                    ),
                    "contrib_2025": (
                        (sales_2025 / summary["2025"]["net_sales"] * 100)
                        if summary["2025"]["net_sales"] > 0
                        else 0.0
                    ),
                }
            )
        return sorted(result, key=lambda item: item["2026"]["net_sales"], reverse=True)

    daily = flatten(by_day)
    for item in daily:
        date_value = datetime.strptime(f"2026-{item['name']}", "%Y-%m-%d")
        item["name"] = f"{date_value.day} {MONTH_NAMES[date_value.strftime('%m')]} 2026"

    def period(year: int) -> str:
        year_dates = sorted(row["date"] for row in filtered if row["year"] == year)
        if not year_dates:
            return f"{year}-01-01 -> {year}-01-01"
        return f"{year_dates[0]} -> {year_dates[-1]}"

    target_daily = read_target_daily(
        month,
        date,
        outlet,
        area,
        available_dates=current_period_dates,
        start_date=start_date,
        end_date=end_date,
    )

    has_date_range = bool(start_date or end_date)
    is_monthly_view = not month and not has_date_range

    if is_monthly_view:
        # YTD / Semua Bulan (tanpa custom date range) -> Tampilkan agregasi per Bulan
        m_keys = sorted({row["month"] for row in filtered if row["year"] == 2026})
        daily_chart = [
            {
                "date": m,
                "comparison_date": f"2025-{m[5:]}",
                "label": MONTH_NAMES.get(m[5:], m[5:]),
                "short_label": datetime.strptime(f"{m}-01", "%Y-%m-%d").strftime("%b"),
                "is_monthly": True,
                "2026": by_month.get(m, {}).get("2026", {"net_sales": 0})["net_sales"],
                "2025": by_month.get(m, {}).get("2025", {"net_sales": 0})["net_sales"],
            }
            for m in m_keys
        ]
        target_chart = [
            {
                "date": item["date"],
                "label": item["label"],
                "short_label": item["short_label"],
                "is_monthly": True,
                "target": sum(
                    v for d_str, v in target_daily.items() if d_str.startswith(item["date"])
                ),
                "actual": item["2026"],
            }
            for item in daily_chart
        ]
    else:
        # Bulan Tertentu atau Rentang Tanggal Dipilih -> Tampilkan agregasi per Hari / Tanggal
        daily_chart = [
            {
                "date": f"2026-{day_month}",
                "comparison_date": f"2025-{day_month}",
                "label": (
                    lambda date_value: f"{date_value.day} {date_value.strftime('%b')}"
                )(datetime.strptime(f"2026-{day_month}", "%Y-%m-%d")),
                "short_label": (
                    lambda date_value: f"{date_value.day} {date_value.strftime('%b')}"
                )(datetime.strptime(f"2026-{day_month}", "%Y-%m-%d")),
                "is_monthly": False,
                "2026": by_day.get(day_month, {}).get("2026", {"net_sales": 0})["net_sales"],
                "2025": by_day.get(day_month, {}).get("2025", {"net_sales": 0})["net_sales"],
            }
            for day_month in sorted(
                {row["date"][5:] for row in filtered if row["year"] == 2026}
            )
        ]
        target_chart = [
            {
                "date": item["date"],
                "label": item["label"],
                "short_label": item["short_label"],
                "is_monthly": False,
                "target": target_daily.get(item["date"], 0.0),
                "actual": item["2026"],
            }
            for item in daily_chart
        ]

    total_target = sum(target_daily.values())
    actual_revenue = summary["2026"]["net_sales"]
    target_achievement = (actual_revenue / total_target * 100) if total_target > 0 else 0.0
    target_gap = actual_revenue - total_target

    # Outlets and products lists
    flat_outlets = flatten(by_outlet)
    flat_products = flatten(by_product)
    flat_areas = flatten(by_area)
    flat_hourly = [item for item in sorted(flatten(by_hour), key=lambda item: int(item["name"])) if int(item["name"]) <= 20]

    # Active outlets with 2026 sales
    active_outlets_2026 = [o for o in flat_outlets if o["2026"]["net_sales"] > 0]
    top_5_outlets = active_outlets_2026[:5]
    bottom_5_outlets = sorted(active_outlets_2026, key=lambda item: item["2026"]["net_sales"])[:5] if len(active_outlets_2026) >= 5 else []
    top_5_products = flat_products[:5]

    # Peak hour analysis
    peak_sales_hour = max(flat_hourly, key=lambda x: x["2026"]["net_sales"])["name"] if flat_hourly else "12"
    peak_tx_hour = max(flat_hourly, key=lambda x: x["2026"].get("transactions", 0))["name"] if flat_hourly else "12"

    # Executive narrative generation
    growth_sales_val = summary["growth"]["net_sales"]
    growth_str = f"{growth_sales_val:+.2f}%" if growth_sales_val is not None else "0.0%"
    target_achieve_str = f"{target_achievement:.1f}%"
    
    status_text = "MELAMPAUI TARGET" if target_gap >= 0 else "DEFISIT DARI TARGET"
    executive_narrative = (
        f"Total penjualan 2026 tercatat sebesar Rp {actual_revenue:,.0f} ({target_achieve_str} dari target Rp {total_target:,.0f}). "
        f"Performa penjualan tercatat {'tumbuh' if (growth_sales_val or 0) >= 0 else 'terkoreksi'} {growth_str} dibanding periode yang sama tahun 2025. "
        f"Rata-rata transaksi per pelanggan (ATV) adalah Rp {summary['2026']['atv']:,.0f} dengan puncak penjualan pada pukul {peak_sales_hour}:00."
    )

    # Visitor Actual Data Integration
    all_raw_visitors = read_all_visitors()
    total_visitors_2026 = 0
    matched_visitor_unit = None
    if outlet and outlet in OUTLET_TO_VISITOR_UNIT:
        matched_visitor_unit = OUTLET_TO_VISITOR_UNIT[outlet]

    for v in all_raw_visitors:
        v_date = v["date"]
        v_dm = v_date[5:]
        if start_date or end_date:
            if start_dm and v_dm < start_dm:
                continue
            if end_dm and v_dm > end_dm:
                continue
        elif month and v["month"] != month:
            continue

        if date and v_date != date:
            continue

        if matched_visitor_unit:
            if v["unit"].casefold() != matched_visitor_unit.casefold():
                continue
        elif area:
            if v["area"].casefold() != area.casefold():
                continue

        total_visitors_2026 += v["visitors"]

    sph_2026 = (actual_revenue / total_visitors_2026) if total_visitors_2026 > 0 else 0.0
    capture_rate = (summary["2026"]["transactions"] / total_visitors_2026 * 100) if total_visitors_2026 > 0 else 0.0

    return {
        "summary": summary,
        "outlets": flat_outlets,
        "top_5_outlets": top_5_outlets,
        "bottom_5_outlets": bottom_5_outlets,
        "products": flat_products,
        "top_5_products": top_5_products,
        "monthly": flatten(by_month),
        "daily": daily,
        "areas": flat_areas,
        "hourly": flat_hourly,
        "hourly_chart": [
            {
                "hour": item["name"],
                "net_sales_2026": item["2026"]["net_sales"],
                "net_sales_2025": item["2025"]["net_sales"],
                "transactions_2026": item["2026"].get("transactions", 0),
                "transactions_2025": item["2025"].get("transactions", 0),
            }
            for item in flat_hourly
        ],
        "target_revenue": total_target,
        "target_achievement": target_achievement,
        "target_gap": target_gap,
        "target_status": status_text,
        "total_visitors": total_visitors_2026,
        "sph": sph_2026,
        "capture_rate": capture_rate,
        "peak_sales_hour": peak_sales_hour,
        "peak_tx_hour": peak_tx_hour,
        "executive_narrative": executive_narrative,
        "months": sorted({row["month"] for row in rows if row["year"] == 2026}, reverse=True),
        "month_labels": {
            month: MONTH_NAMES.get(month[5:], month)
            for month in {row["month"] for row in rows if row["year"] == 2026}
        },
        "dates": sorted({row["date"] for row in rows if row["year"] == 2026}),
        "date_labels": {
            date: (
                lambda date_value: (
                    f"{WEEKDAY_NAMES[date_value.weekday()]}, "
                    f"{date_value.day} {MONTH_NAMES[date_value.strftime('%m')]} 2026"
                )
            )(datetime.strptime(date, "%Y-%m-%d"))
            for date in {row["date"] for row in rows if row["year"] == 2026}
        },
        "outlet_options": sorted(
            {
                row["outlet"]
                for row in rows
                if not area or row["area"] == area
            }
        ),
        "area_options": sorted({row["area"] for row in rows}),
        "total_outlets": len({row["outlet"] for row in filtered}),
        "daily_chart": daily_chart,
        "target_chart": target_chart,
        "comparison_period": {
            "current": period(2026),
            "comparison": period(2025),
        },
        "rows_read": len(rows),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "compact_sales": [
            [
                r["year"],
                r["date"],
                r["hour"],
                r["outlet"],
                r["area"],
                r["product"],
                r["qty"],
                r["net_sales"],
                r.get("invoice", ""),
                r.get("item_net_sales", 0.0),
            ]
            for r in rows
            if r["year"] == 2026 or r["date"][5:] in all_sales_day_months_2026
        ]
        if include_raw
        else [],
        # The browser-side filter engine receives only target days for which
        # 2026 sales data is currently available. This keeps its default and
        # filtered target totals aligned with the backend's MTD calculation.
        "compact_targets": (
            [
                target
                for target in read_all_targets()
                if target["date"] in all_sales_dates_2026
            ]
            if include_raw
            else []
        ),
        "compact_visitors": (
            all_raw_visitors
            if include_raw
            else []
        ),
    }


def build_excel_report(data: dict[str, Any], filter_desc: str = "") -> BytesIO:
    wb = Workbook()

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="Segoe UI", size=14, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    regular_font = Font(name="Segoe UI", size=10)

    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # 1. Sheet: Ringkasan KPI
    ws_sum = wb.active
    ws_sum.title = "Ringkasan KPI"
    ws_sum.views.sheetView[0].showGridLines = True

    ws_sum["A1"] = "SALES PERFORMANCE & GROWTH DASHBOARD"
    ws_sum["A1"].font = title_font
    ws_sum["A2"] = f"Filter: {filter_desc or 'Semua Data'} | Diunduh: {datetime.now().strftime('%d %b %Y %H:%M:%S')}"
    ws_sum["A2"].font = subtitle_font

    ws_sum["A4"] = "RINGKASAN EKSEKUTIF:"
    ws_sum["A4"].font = bold_font
    ws_sum["A5"] = data.get("executive_narrative", "")
    ws_sum["A5"].font = regular_font

    headers = [
        "Indikator / Metrik",
        "Tahun 2026 (Utama)",
        "Tahun 2025 (Pembanding)",
        "Pertumbuhan YoY (%)",
        "Selisih (+/-)",
    ]
    for col_num, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=7, column=col_num, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    kpi_rows = [
        (
            "Net Sales (Omset Bersih)",
            data["summary"]["2026"]["net_sales"],
            data["summary"]["2025"]["net_sales"],
            data["summary"]["growth"]["net_sales"],
            data["summary"]["diff"]["net_sales"],
            "currency",
        ),
        (
            "Target Revenue 2026",
            data["target_revenue"],
            None,
            None,
            data["target_gap"],
            "currency",
        ),
        (
            "Pencapaian Target (%)",
            data["target_achievement"] / 100 if data["target_revenue"] else 0,
            None,
            None,
            None,
            "percent",
        ),
        (
            "Total Transaksi (Struk)",
            data["summary"]["2026"]["transactions"],
            data["summary"]["2025"]["transactions"],
            data["summary"]["growth"]["transactions"],
            data["summary"]["diff"]["transactions"],
            "number",
        ),
        (
            "Total Qty Terjual (Pcs)",
            data["summary"]["2026"]["qty"],
            data["summary"]["2025"]["qty"],
            data["summary"]["growth"]["qty"],
            data["summary"]["diff"]["qty"],
            "number",
        ),
        (
            "ATV (Average Transaction Value)",
            data["summary"]["2026"]["atv"],
            data["summary"]["2025"]["atv"],
            data["summary"]["growth"]["atv"],
            data["summary"]["diff"]["atv"],
            "currency",
        ),
        (
            "UPT (Units Per Transaction)",
            data["summary"]["2026"]["upt"],
            data["summary"]["2025"]["upt"],
            data["summary"]["growth"]["upt"],
            None,
            "decimal",
        ),
        (
            "ASP (Average Selling Price)",
            data["summary"]["2026"]["asp"],
            data["summary"]["2025"]["asp"],
            data["summary"]["growth"]["asp"],
            None,
            "currency",
        ),
    ]

    for r_idx, (label, v26, v25, growth, diff, fmt) in enumerate(kpi_rows, 8):
        c_lbl = ws_sum.cell(row=r_idx, column=1, value=label)
        c_lbl.font = bold_font
        c26 = ws_sum.cell(row=r_idx, column=2, value=v26 if v26 is not None else "-")
        c25 = ws_sum.cell(row=r_idx, column=3, value=v25 if v25 is not None else "-")
        cg = ws_sum.cell(row=r_idx, column=4, value=(growth / 100) if growth is not None else "-")
        cd = ws_sum.cell(row=r_idx, column=5, value=diff if diff is not None else "-")

        for c in (c_lbl, c26, c25, cg, cd):
            c.border = thin_border
            c.font = regular_font if c != c_lbl else bold_font

        if fmt == "currency":
            if isinstance(v26, (int, float)): c26.number_format = "#,##0"
            if isinstance(v25, (int, float)): c25.number_format = "#,##0"
            if isinstance(diff, (int, float)): cd.number_format = "+#,##0;-#,##0;0"
        elif fmt == "number":
            if isinstance(v26, (int, float)): c26.number_format = "#,##0"
            if isinstance(v25, (int, float)): c25.number_format = "#,##0"
            if isinstance(diff, (int, float)): cd.number_format = "+#,##0;-#,##0;0"
        elif fmt == "percent":
            if isinstance(v26, (int, float)): c26.number_format = "0.00%"
        elif fmt == "decimal":
            if isinstance(v26, (int, float)): c26.number_format = "0.00"
            if isinstance(v25, (int, float)): c25.number_format = "0.00"

        if isinstance(growth, (int, float)):
            cg.number_format = "+0.00%;-0.00%;0.00%"

    # 2. Sheet: Performa Outlet
    ws_out = wb.create_sheet(title="Performa Outlet")
    ws_out.views.sheetView[0].showGridLines = True
    out_headers = [
        "Rank",
        "Nama Konter / Outlet",
        "Net Sales 2026",
        "Kontribusi 2026 (%)",
        "Qty 2026",
        "Transaksi 2026",
        "Net Sales 2025",
        "Growth Sales (%)",
    ]
    for col_num, h in enumerate(out_headers, 1):
        cell = ws_out.cell(row=1, column=col_num, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, o in enumerate(data.get("outlets", []), 2):
        ws_out.cell(row=r_idx, column=1, value=r_idx - 1).alignment = Alignment(horizontal="center")
        ws_out.cell(row=r_idx, column=2, value=o["name"]).font = bold_font
        c_s26 = ws_out.cell(row=r_idx, column=3, value=o["2026"]["net_sales"])
        c_s26.number_format = "#,##0"
        c_cb = ws_out.cell(row=r_idx, column=4, value=(o.get("contrib_2026", 0) / 100))
        c_cb.number_format = "0.00%"
        c_q26 = ws_out.cell(row=r_idx, column=5, value=o["2026"]["qty"])
        c_q26.number_format = "#,##0"
        c_t26 = ws_out.cell(row=r_idx, column=6, value=o["2026"].get("transactions", 0))
        c_t26.number_format = "#,##0"
        c_s25 = ws_out.cell(row=r_idx, column=7, value=o["2025"]["net_sales"])
        c_s25.number_format = "#,##0"
        gw = o.get("growth_sales")
        c_gw = ws_out.cell(row=r_idx, column=8, value=(gw / 100) if gw is not None else "-")
        if gw is not None:
            c_gw.number_format = "+0.00%;-0.00%;0.00%"
        for col_idx in range(1, 9):
            ws_out.cell(row=r_idx, column=col_idx).border = thin_border
            ws_out.cell(row=r_idx, column=col_idx).font = regular_font if col_idx != 2 else bold_font

    # 3. Sheet: Performa Area
    ws_ar = wb.create_sheet(title="Performa Area")
    ws_ar.views.sheetView[0].showGridLines = True
    for col_num, h in enumerate(
        ["Rank", "Nama Area", "Net Sales 2026", "Kontribusi 2026 (%)", "Qty 2026", "Transaksi 2026", "Net Sales 2025", "Growth Sales (%)"],
        1,
    ):
        cell = ws_ar.cell(row=1, column=col_num, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, a in enumerate(data.get("areas", []), 2):
        ws_ar.cell(row=r_idx, column=1, value=r_idx - 1).alignment = Alignment(horizontal="center")
        ws_ar.cell(row=r_idx, column=2, value=a["name"]).font = bold_font
        c_s26 = ws_ar.cell(row=r_idx, column=3, value=a["2026"]["net_sales"])
        c_s26.number_format = "#,##0"
        c_cb = ws_ar.cell(row=r_idx, column=4, value=(a.get("contrib_2026", 0) / 100))
        c_cb.number_format = "0.00%"
        c_q26 = ws_ar.cell(row=r_idx, column=5, value=a["2026"]["qty"])
        c_q26.number_format = "#,##0"
        c_t26 = ws_ar.cell(row=r_idx, column=6, value=a["2026"].get("transactions", 0))
        c_t26.number_format = "#,##0"
        c_s25 = ws_ar.cell(row=r_idx, column=7, value=a["2025"]["net_sales"])
        c_s25.number_format = "#,##0"
        gw = a.get("growth_sales")
        c_gw = ws_ar.cell(row=r_idx, column=8, value=(gw / 100) if gw is not None else "-")
        if gw is not None:
            c_gw.number_format = "+0.00%;-0.00%;0.00%"
        for col_idx in range(1, 9):
            ws_ar.cell(row=r_idx, column=col_idx).border = thin_border
            ws_ar.cell(row=r_idx, column=col_idx).font = regular_font if col_idx != 2 else bold_font

    # 4. Sheet: Top Produk
    ws_pr = wb.create_sheet(title="Top Produk")
    ws_pr.views.sheetView[0].showGridLines = True
    for col_num, h in enumerate(
        ["Rank", "Nama Barang / Produk", "Net Sales 2026", "Kontribusi 2026 (%)", "Qty 2026", "Qty 2025", "Growth Qty (%)"],
        1,
    ):
        cell = ws_pr.cell(row=1, column=col_num, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, p in enumerate(data.get("products", [])[:200], 2):
        ws_pr.cell(row=r_idx, column=1, value=r_idx - 1).alignment = Alignment(horizontal="center")
        ws_pr.cell(row=r_idx, column=2, value=p["name"]).font = bold_font
        c_s26 = ws_pr.cell(row=r_idx, column=3, value=p["2026"]["net_sales"])
        c_s26.number_format = "#,##0"
        c_cb = ws_pr.cell(row=r_idx, column=4, value=(p.get("contrib_2026", 0) / 100))
        c_cb.number_format = "0.00%"
        c_q26 = ws_pr.cell(row=r_idx, column=5, value=p["2026"]["qty"])
        c_q26.number_format = "#,##0"
        c_q25 = ws_pr.cell(row=r_idx, column=6, value=p["2025"]["qty"])
        c_q25.number_format = "#,##0"
        gq = p.get("growth_qty")
        c_gq = ws_pr.cell(row=r_idx, column=7, value=(gq / 100) if gq is not None else "-")
        if gq is not None:
            c_gq.number_format = "+0.00%;-0.00%;0.00%"
        for col_idx in range(1, 8):
            ws_pr.cell(row=r_idx, column=col_idx).border = thin_border
            ws_pr.cell(row=r_idx, column=col_idx).font = regular_font if col_idx != 2 else bold_font

    # 5. Sheet: Performa Jam
    ws_hr = wb.create_sheet(title="Performa Jam")
    ws_hr.views.sheetView[0].showGridLines = True
    for col_num, h in enumerate(
        ["Jam Operasional", "Net Sales 2026", "Transaksi 2026", "Net Sales 2025", "Transaksi 2025"],
        1,
    ):
        cell = ws_hr.cell(row=1, column=col_num, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, hr in enumerate(data.get("hourly_chart", []), 2):
        ws_hr.cell(row=r_idx, column=1, value=f"Pukul {hr['hour']}:00").alignment = Alignment(horizontal="center")
        c_s26 = ws_hr.cell(row=r_idx, column=2, value=hr["net_sales_2026"])
        c_s26.number_format = "#,##0"
        c_t26 = ws_hr.cell(row=r_idx, column=3, value=hr["transactions_2026"])
        c_t26.number_format = "#,##0"
        c_s25 = ws_hr.cell(row=r_idx, column=4, value=hr["net_sales_2025"])
        c_s25.number_format = "#,##0"
        c_t25 = ws_hr.cell(row=r_idx, column=5, value=hr["transactions_2025"])
        c_t25.number_format = "#,##0"
        for col_idx in range(1, 6):
            ws_hr.cell(row=r_idx, column=col_idx).border = thin_border
            ws_hr.cell(row=r_idx, column=col_idx).font = regular_font

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def write_report(html: str) -> None:
    """Optionally persist a static snapshot.

    The interactive dashboard includes its source data for client-side filters,
    so saving every rendered page produced 90 MB HTML files. Static snapshots
    are therefore opt-in via WRITE_STATIC_REPORT=1.
    """
    if os.environ.get("WRITE_STATIC_REPORT") != "1":
        return
    REPORT_DIR.mkdir(exist_ok=True)
    (REPORT_DIR / "sales_dashboard.html").write_text(html, encoding="utf-8")


@app.after_request
def compress_large_response(response):
    """Reduce transfer size for the data-heavy legacy client without UI changes."""
    accepts_gzip = "gzip" in request.headers.get("Accept-Encoding", "").lower()
    compressible = response.mimetype in {"text/html", "application/json"}
    if (
        response.direct_passthrough
        or not accepts_gzip
        or not compressible
        or response.headers.get("Content-Encoding")
    ):
        return response

    payload = response.get_data()
    if len(payload) < 1024:
        return response
    response.set_data(gzip.compress(payload, compresslevel=6))
    response.headers["Content-Encoding"] = "gzip"
    response.headers["Vary"] = "Accept-Encoding"
    return response


@app.route("/")
def index():
    month = request.args.get("month") or None
    date = request.args.get("date") or None
    start_date = request.args.get("start_date") or None
    end_date = request.args.get("end_date") or None
    outlet = request.args.get("outlet") or None
    area = request.args.get("area") or None
    html = render_template(
        "index.html",
        data=build_dashboard(
            read_sales(),
            month,
            date,
            outlet,
            area,
            start_date=start_date,
            end_date=end_date,
        ),
        selected_month=month or "",
        selected_date=date or "",
        selected_outlet=outlet or "",
        selected_area=area or "",
    )
    write_report(html)
    return html


@app.route("/download")
def download_excel():
    month = request.args.get("month") or None
    date = request.args.get("date") or None
    start_date = request.args.get("start_date") or None
    end_date = request.args.get("end_date") or None
    outlet = request.args.get("outlet") or None
    area = request.args.get("area") or None

    all_rows = read_sales()
    dashboard_data = build_dashboard(
        all_rows,
        month,
        date,
        outlet,
        area,
        include_raw=False,
        start_date=start_date,
        end_date=end_date,
    )

    filter_parts = []
    if month: filter_parts.append(f"Bulan_{month}")
    if start_date or end_date:
        filter_parts.append(f"Tgl_{start_date or 'Awal'}_sd_{end_date or 'Akhir'}")
    elif date:
        filter_parts.append(f"Tgl_{date}")
    if area: filter_parts.append(f"Area_{area}")
    if outlet: filter_parts.append(f"Outlet_{outlet}")

    filter_desc = " - ".join(filter_parts) if filter_parts else "Semua Data"
    file_name = f"Laporan_Sales_{'_'.join(filter_parts) if filter_parts else 'Semua'}.xlsx"

    excel_buffer = build_excel_report(dashboard_data, filter_desc)
    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/dashboard")
def dashboard_api():
    return jsonify(
        build_dashboard(
            read_sales(),
            request.args.get("month") or None,
            request.args.get("date") or None,
            request.args.get("outlet") or None,
            request.args.get("area") or None,
            include_raw=False,
            start_date=request.args.get("start_date") or None,
            end_date=request.args.get("end_date") or None,
        )
    )


# ─────────────────────────────────────────────────────────────────────────────
# ADMIN: Upload File Excel Sales & Target
# ─────────────────────────────────────────────────────────────────────────────
ADMIN_PIN = os.environ.get("ADMIN_PIN", "1234")
ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
app.config.setdefault("MAX_CONTENT_LENGTH", 64 * 1024 * 1024)  # 64 MB


def _allowed_file(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def _invalidate_sales_cache() -> None:
    global _sales_cache, _sales_cache_signature, _sales_cache_mapping_signature
    global _budget_cache, _budget_cache_signature
    with _sales_cache_lock:
        _sales_cache = None
        _sales_cache_signature = ()
        _sales_cache_mapping_signature = None
    with _budget_cache_lock:
        _budget_cache = None
        _budget_cache_signature = None


@app.route("/upload", methods=["GET", "POST"])
def upload_file():
    """
    GET  -> tampilkan halaman upload admin
    POST -> terima file Excel, validasi PIN, simpan ke folder yang sesuai,
           lalu invalidate cache agar data langsung ter-refresh.
    """
    if request.method == "GET":
        return render_template("upload.html", message=None, error=None)

    # Validasi PIN
    pin = (request.form.get("pin") or "").strip()
    if pin != ADMIN_PIN:
        return render_template("upload.html", message=None, error="❌ PIN salah. Akses ditolak."), 403

    # Validasi file yang di-upload
    files = request.files.getlist("files")
    if not files or all(f.filename == "" for f in files):
        return render_template("upload.html", message=None, error="❌ Tidak ada file yang dipilih."), 400

    saved_files: list[str] = []
    skipped_files: list[str] = []

    for f in files:
        if not f or f.filename == "":
            continue
        fname = secure_filename(f.filename)
        if not _allowed_file(fname):
            skipped_files.append(f"{fname} (bukan .xlsx/.xls)")
            continue

        category = request.form.get("category", "auto")
        raw_name = f.filename.lower()
        if category == "2025":
            dest_dir = SALES_ROOT_DIR / "sales detail 2025"
        elif category == "2026":
            dest_dir = SALES_ROOT_DIR / "sales detail 2026"
        elif category == "target":
            dest_dir = BUDGET_FILE.parent
        else:
            if "target" in raw_name or "budget" in raw_name:
                dest_dir = BUDGET_FILE.parent
            elif "2025" in raw_name:
                dest_dir = SALES_ROOT_DIR / "sales detail 2025"
            else:
                dest_dir = SALES_ROOT_DIR / "sales detail 2026"

        dest_dir.mkdir(parents=True, exist_ok=True)
        dest_path = dest_dir / fname

        f.save(str(dest_path))
        saved_files.append(f"{fname} → {dest_dir.name}/")

    if not saved_files:
        return render_template(
            "upload.html",
            message=None,
            error="❌ Tidak ada file valid yang berhasil disimpan. " + (", ".join(skipped_files) if skipped_files else ""),
        ), 400

    # Invalidate cache
    _invalidate_sales_cache()

    # Re-render static report
    try:
        updated_rows = read_sales()
        dashboard_data = build_dashboard(updated_rows)
        with app.app_context():
            write_report(
                render_template(
                    "index.html",
                    data=dashboard_data,
                    selected_month="",
                    selected_date="",
                    selected_outlet="",
                    selected_area="",
                )
            )
    except Exception:
        pass

    msg_parts = [f"✅ {len(saved_files)} file berhasil disimpan:"]
    for s in saved_files:
        msg_parts.append(f"  • {s}")
    if skipped_files:
        msg_parts.append(f"⚠️ Dilewati: {', '.join(skipped_files)}")
    msg_parts.append("🔄 Cache data di-reset. Dashboard sudah ter-refresh otomatis!")

    return render_template("upload.html", message="\n".join(msg_parts), error=None)


@app.route("/api/clear-cache", methods=["POST"])
def api_clear_cache():
    body = request.get_json(silent=True) or {}
    if str(body.get("pin", "")).strip() != ADMIN_PIN:
        return jsonify({"ok": False, "error": "PIN salah"}), 403
    _invalidate_sales_cache()
    return jsonify({"ok": True, "message": "Cache di-reset. Data akan di-reload pada request berikutnya."})


@app.route("/api/save-visitor", methods=["POST"])
def api_save_visitor():
    """Endpoint untuk input/update data pengunjung harian per unit wahana."""
    body = request.get_json(silent=True)
    if not body:
        # Fallback to form-data if submitted via regular form
        body = request.form

    pin = str(body.get("pin", "")).strip()
    if pin != ADMIN_PIN:
        return jsonify({"ok": False, "error": "PIN Admin salah. Akses ditolak."}), 403

    entry_date = str(body.get("date", "")).strip()
    unit = str(body.get("unit", "")).strip()
    count_val = body.get("count", 0)

    if not entry_date:
        return jsonify({"ok": False, "error": "Tanggal harus diisi."}), 400
    if not unit:
        return jsonify({"ok": False, "error": "Unit wahana harus dipilih."}), 400

    try:
        count_int = int(count_val)
        if count_int < 0:
            return jsonify({"ok": False, "error": "Jumlah pengunjung tidak boleh negatif."}), 400
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "Format jumlah pengunjung harus berupa angka bulat."}), 400

    try:
        success = save_visitor_actual(entry_date, unit, count_int)
        if not success:
            return jsonify({"ok": False, "error": "Gagal menyimpan ke file budget."}), 500

        # Re-render static report
        try:
            updated_rows = read_sales()
            dashboard_data = build_dashboard(updated_rows)
            with app.app_context():
                write_report(
                    render_template(
                        "index.html",
                        data=dashboard_data,
                        selected_month="",
                        selected_date="",
                        selected_outlet="",
                        selected_area="",
                    )
                )
        except Exception:
            pass

        return jsonify({
            "ok": True,
            "message": f"✅ Berhasil menyimpan {count_int:,} pengunjung untuk {unit} pada tanggal {entry_date}."
        })
    except Exception as e:
        return jsonify({"ok": False, "error": f"Terjadi kesalahan: {str(e)}"}), 500




if __name__ == "__main__":
    initial_data = build_dashboard(read_sales())
    with app.app_context():
        write_report(
            render_template(
                "index.html",
                data=initial_data,
                selected_month="",
                selected_date="",
                selected_outlet="",
                selected_area="",
            )
        )
    app.run(debug=False, use_reloader=False, host="127.0.0.1", port=5000)
