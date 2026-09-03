import math
import os
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

# Fix Windows console encoding for emoji characters
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

from openpyxl import load_workbook

from database.db_config import BASE_DIR, get_connection, init_db

SALES_ROOT_DIR = BASE_DIR / "data"
SALES_DETAIL_DIR_PATTERN = re.compile(r"^sales\s+detail\s+(\d{4})$", re.I)
MAPPING_FILE = BASE_DIR / "config" / "MASTER_OUTLET_MAPPING_V2.xlsx"
BUDGET_FILE = BASE_DIR / "data" / "target 2026" / "BUDGET_MERCH_ONLY_PYTHON_READY.xlsx"

SALES_VALUE_COLUMN_INDEX = 8         # Column I (DPP Barang)
TRANSACTION_TOTAL_COLUMN_INDEX = 17  # Column R (Total Struk)
TRANSACTION_TOTAL_TAX_FACTOR = 1.11
INVOICE_DISCOUNT_COLUMN_INDEX = 13   # Column N (Diskon Faktur)

VISITOR_UNIT_TO_AREA = {
    "SEAWORLD": "AWAPARK",
    "SAMUDRA": "AWAPARK",
    "ATLANTIS": "AWAPARK",
    "BEACHPARK": "BEACHPARK",
    "DUFAN": "DUFAN",
}


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value) if math.isfinite(float(value)) else 0.0
    cleaned = re.sub(r"[^\d,.-]", "", str(value).strip())
    if not cleaned:
        return 0.0
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    for fmt in ("%d %b %Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return None


MAPPING_FILE_CANDIDATES = [
    BASE_DIR / "config" / "MASTER_OUTLET_MAPPING_V2.xlsx",
    BASE_DIR / "config" / "MASTER_OUTLET_MAPPING.xlsx",
    BASE_DIR / "Porto" / "docs" / "MASTER_OUTLET_MAPPING.xlsx",
    BASE_DIR / "docs" / "MASTER_OUTLET_MAPPING.xlsx",
]

BUDGET_FILE_CANDIDATES = [
    BASE_DIR / "data" / "target 2026" / "BUDGET_MERCH_ONLY_PYTHON_READY.xlsx",
    BASE_DIR / "data" / "target 2026" / "MASTER_TARGET_VISITOR_TEMPLATE_V3.xlsx",
    BASE_DIR / "Porto" / "docs" / "MASTER_TARGET_VISITOR_TEMPLATE_V3.xlsx",
    BASE_DIR / "docs" / "MASTER_TARGET_VISITOR_TEMPLATE_V3.xlsx",
]


def _get_active_mapping_file() -> Path | None:
    for p in MAPPING_FILE_CANDIDATES:
        if p.exists():
            return p
    return None


def _get_active_budget_file() -> Path | None:
    for p in BUDGET_FILE_CANDIDATES:
        if p.exists():
            return p
    return None


def read_outlet_mapping() -> dict[str, str]:
    mapping: dict[str, str] = {
        # Default Known Ancol Retail Outlets Fallback
        "DFIN Dufan Induk": "DUFAN",
        "DFIL Dufan Induk Lama": "DUFAN",
        "DFKE Dufan Kereta Misteri": "DUFAN",
        "DFAR Dufan Arung Jeram": "DUFAN",
        "DFTO Dufan Tornado": "DUFAN",
        "DFHL Dufan Halilintar": "DUFAN",
        "DFOP Dufan Ontang Anting": "DUFAN",
        "DFHY Dufan Hysteria": "DUFAN",
        "DFBI Dufan Bianglala": "DUFAN",
        "DFIC Dufan Ice Cream": "DUFAN",
        "SWIN Sea World Induk": "AWAPARK",
        "SWTG Sea World Touchpool": "AWAPARK",
        "ODIN Samudra Induk": "AWAPARK",
        "ATIN Atlantis Induk": "AWAPARK",
        "AWKL AWA Taman Kelapa 2": "AWAPARK",
        "TJOM Merchandise Ombak Laut": "BEACHPARK",
        "JBIN JBL Induk": "BEACHPARK",
        "OL01 Online Shop": "BEACHPARK",
    }
    
    mapping_path = _get_active_mapping_file()
    if not mapping_path:
        return mapping

    try:
        wb = load_workbook(mapping_path, read_only=True, data_only=True)
        sheet = wb["OUTLET_MAPPING"] if "OUTLET_MAPPING" in wb.sheetnames else wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2 and row[0] and row[1]:
                outlet_name = str(row[0]).strip()
                area_name = str(row[1]).strip().upper()
                mapping[outlet_name] = area_name
        wb.close()
    except Exception as e:
        print(f"⚠️ Warning saat membaca outlet mapping: {e}")
    
    return mapping


def _budget_day_type(day: datetime, holidays: dict[str, tuple[str, float | None]]) -> tuple[str, float | None]:
    holiday = holidays.get(day.strftime("%Y-%m-%d"))
    if holiday:
        return holiday
    if day.weekday() == 5:
        return "SATURDAY", None
    if day.weekday() == 6:
        return "SUNDAY", None
    return "WEEKDAY", None


def _budget_outlet_name(code: Any, name: Any, outlet_mapping: dict[str, str]) -> str:
    code_text = str(code).strip() if code else ""
    name_text = str(name).strip() if name else ""
    prefix = f"{code_text} "
    matched = next((outlet for outlet in outlet_mapping if outlet.startswith(prefix)), None)
    return matched or f"{code_text} {name_text}".strip()


def parse_sales_file(path: Path, outlet_mapping: dict[str, str]) -> list[dict[str, Any]]:
    """Parse single sales detail Excel file using transaction-level Net Sales allocation."""
    year_match = SALES_DETAIL_DIR_PATTERN.match(path.parent.name)
    file_year = int(year_match.group(1)) if year_match else None
    if file_year not in (2025, 2026, 2027, 2028):
        return []

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"⚠️ Gagal membuka file {path.name}: {e}")
        return []

    rows: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            outlet: str | None = None
            header_seen = False
            current_invoice: str | None = None
            current_date: datetime | None = None
            current_hour = "00"
            current_items: list[dict[str, Any]] = []

            def finalize_transaction(total_value: Any, invoice_discount: Any = 0) -> None:
                nonlocal current_items
                if not current_items:
                    return

                total_struk = _number(total_value)
                discount_invoice = _number(invoice_discount)
                transaction_net_sales = total_struk / TRANSACTION_TOTAL_TAX_FACTOR
                item_i_total = sum(item["item_net_sales"] for item in current_items)

                if item_i_total != 0:
                    for item in current_items:
                        item["net_sales"] = (
                            transaction_net_sales
                            * item["item_net_sales"]
                            / item_i_total
                        )
                else:
                    share = (
                        transaction_net_sales / len(current_items)
                        if current_items
                        else 0.0
                    )
                    for item in current_items:
                        item["net_sales"] = share

                for item in current_items:
                    item["transaction_total"] = total_struk
                    item["invoice_discount"] = discount_invoice

                rows.extend(current_items)
                current_items = []

            for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = list(values)
                first = str(cells[0]).strip() if cells and cells[0] is not None else ""

                if (
                    first
                    and first not in {
                        "Penjualan Cabang (POS)",
                        "Tanggal :",
                        "Kasir :",
                        "No. Invoice",
                        "Kode Barang",
                        "Total Struk",
                    }
                    and all(value is None for value in cells[1:7])
                ):
                    if current_items:
                        finalize_transaction(None)
                    outlet = first
                    header_seen = False
                    current_invoice = None
                    current_date = None
                    current_hour = "00"
                    continue

                if first == "No. Invoice":
                    header_seen = True
                    continue

                if not header_seen or not outlet:
                    continue

                if first.startswith("Total "):
                    if first == "Total Struk":
                        total_struk_val = (
                            cells[TRANSACTION_TOTAL_COLUMN_INDEX]
                            if len(cells) > TRANSACTION_TOTAL_COLUMN_INDEX
                            else None
                        )
                        disc_val = (
                            cells[INVOICE_DISCOUNT_COLUMN_INDEX]
                            if len(cells) > INVOICE_DISCOUNT_COLUMN_INDEX
                            else 0
                        )
                        finalize_transaction(total_struk_val, disc_val)
                    continue

                if len(cells) > 2 and cells[2] is not None:
                    parsed_date = _date(cells[2])
                    if parsed_date:
                        current_date = parsed_date

                if len(cells) > 4 and cells[4] is not None:
                    time_text = str(cells[4]).strip()
                    if ":" in time_text:
                        current_hour = time_text.split(":")[0].zfill(2)

                if first and not first.startswith("Total "):
                    current_invoice = first

                product = str(cells[1]).strip() if len(cells) > 1 and cells[1] is not None else ""
                if not product:
                    continue

                qty = _number(cells[6]) if len(cells) > 6 else 0.0
                sales_raw = (
                    _number(cells[SALES_VALUE_COLUMN_INDEX])
                    if len(cells) > SALES_VALUE_COLUMN_INDEX
                    else 0.0
                )

                if current_date is None:
                    continue

                date_str = current_date.strftime("%Y-%m-%d")
                row_year = current_date.year
                if row_year != file_year:
                    continue

                area = outlet_mapping.get(outlet, "UNKNOWN")
                current_items.append({
                    "file_source": path.name,
                    "year": row_year,
                    "date": date_str,
                    "month": current_date.strftime("%Y-%m"),
                    "hour": current_hour,
                    "invoice": current_invoice or "",
                    "outlet": outlet,
                    "area": area,
                    "product": product,
                    "qty": qty,
                    "item_net_sales": sales_raw,
                    "net_sales": sales_raw,
                    "transaction_total": 0.0,
                    "invoice_discount": 0.0,
                })

            if current_items:
                finalize_transaction(None)
    finally:
        workbook.close()

    return rows


def parse_budget_and_visitor(budget_path: Path, mapping_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Parse master budget file for daily target rows and visitor actuals."""
    if not budget_path.exists():
        return [], []

    outlet_mapping = read_outlet_mapping()
    wb = load_workbook(budget_path, read_only=True, data_only=True)
    daily_targets: list[dict[str, Any]] = []
    visitors: list[dict[str, Any]] = []

    try:
        # 1. Budget Targets dari sheet TARGET_REVENUE, DAY_WEIGHTS, dan HOLIDAYS
        if "TARGET_REVENUE" in wb.sheetnames and "DAY_WEIGHTS" in wb.sheetnames:
            weights: dict[str, float] = {}
            for day_type, weight, *_ in wb["DAY_WEIGHTS"].iter_rows(min_row=4, values_only=True):
                if day_type:
                    weights[str(day_type).strip().upper()] = _number(weight)

            holidays: dict[str, tuple[str, float | None]] = {}
            if "HOLIDAYS" in wb.sheetnames:
                for holiday_date, _, day_type, weight_override in wb["HOLIDAYS"].iter_rows(min_row=4, values_only=True):
                    parsed_date = _date(holiday_date)
                    if parsed_date:
                        override = _number(weight_override) if weight_override not in (None, "") else None
                        holidays[parsed_date.strftime("%Y-%m-%d")] = (
                            str(day_type or "HOLIDAY").strip().upper(),
                            override,
                        )

            monthly_rows: list[tuple[datetime, Any, Any, float]] = []
            for month_value, outlet_code, outlet_name, monthly_target in wb["TARGET_REVENUE"].iter_rows(min_row=4, values_only=True):
                parsed_month = _date(month_value)
                if parsed_month and outlet_code:
                    monthly_rows.append((parsed_month, outlet_code, outlet_name, _number(monthly_target)))

            from calendar import monthrange
            for month_value, outlet_code, outlet_name, monthly_target in monthly_rows:
                days_in_month = monthrange(month_value.year, month_value.month)[1]
                daily_plan: list[tuple[datetime, str, float]] = []
                for day_number in range(1, days_in_month + 1):
                    current_day = datetime(month_value.year, month_value.month, day_number)
                    day_type, override = _budget_day_type(current_day, holidays)
                    daily_plan.append((current_day, day_type, override if override is not None else weights.get(day_type, 1.0)))

                total_weight = sum(w for _, _, w in daily_plan)
                resolved_outlet = _budget_outlet_name(outlet_code, outlet_name, outlet_mapping)
                for current_day, day_type, weight in daily_plan:
                    daily_targets.append({
                        "date": current_day.strftime("%Y-%m-%d"),
                        "month": current_day.strftime("%Y-%m"),
                        "outlet": resolved_outlet,
                        "area": outlet_mapping.get(resolved_outlet, "Tidak terpetakan"),
                        "target": monthly_target * weight / total_weight if total_weight else 0.0,
                    })

        # 2. Visitor Actuals
        if "VISITOR_ACTUAL" in wb.sheetnames:
            ws_vis = wb["VISITOR_ACTUAL"]
            for row in ws_vis.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                dt_val = _date(row[0])
                if not dt_val:
                    continue
                unit_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                visitor_count = _number(row[2]) if len(row) > 2 else 0.0
                area_name = VISITOR_UNIT_TO_AREA.get(unit_name.upper(), "")
                visitors.append({
                    "date": dt_val.strftime("%Y-%m-%d"),
                    "month": dt_val.strftime("%Y-%m"),
                    "unit": unit_name,
                    "area": area_name,
                    "visitors": int(visitor_count),
                })
    finally:
        wb.close()

    return daily_targets, visitors


def sync_database(force: bool = False) -> dict[str, Any]:
    """
    Sinkronisasi file Excel ke SQLite Database secara otomatis & bertahap (incremental).
    Hanya file yang baru atau memiliki perubahan mtime/size yang akan diproses.
    """
    init_db()
    conn = get_connection()
    stats = {"sales_files_synced": 0, "sales_rows_inserted": 0, "budget_synced": False, "visitors_count": 0}

    try:
        # Cari semua file sales
        sales_files: list[Path] = []
        if SALES_ROOT_DIR.exists():
            for source_dir in sorted(SALES_ROOT_DIR.iterdir()):
                if not source_dir.is_dir() or not SALES_DETAIL_DIR_PATTERN.match(source_dir.name):
                    continue
                sales_files.extend(
                    p for p in source_dir.glob("*.xlsx") if not p.name.startswith("~$")
                )

        outlet_mapping = read_outlet_mapping()

        # 1. Sync Sales Files
        for s_file in sales_files:
            file_stat = s_file.stat()
            file_key = str(s_file.resolve())
            mtime = int(file_stat.st_mtime_ns)
            size = int(file_stat.st_size)

            cursor = conn.execute(
                "SELECT file_mtime, file_size FROM sync_meta WHERE file_path = ?", (file_key,)
            )
            meta = cursor.fetchone()

            if force or not meta or meta["file_mtime"] != mtime or meta["file_size"] != size:
                print(f"🔄 Sinkronisasi file sales ke DB: {s_file.name} ...")
                rows = parse_sales_file(s_file, outlet_mapping)

                conn.execute("BEGIN TRANSACTION;")
                # Hapus data lama dari file ini jika ada
                conn.execute("DELETE FROM sales_items WHERE file_source = ?", (s_file.name,))
                
                # Batch insert
                conn.executemany(
                    """
                    INSERT INTO sales_items (
                        file_source, year, date, month, hour, invoice,
                        outlet, area, product, qty, item_net_sales,
                        net_sales, transaction_total, invoice_discount
                    ) VALUES (
                        :file_source, :year, :date, :month, :hour, :invoice,
                        :outlet, :area, :product, :qty, :item_net_sales,
                        :net_sales, :transaction_total, :invoice_discount
                    )
                    """,
                    rows,
                )

                # Update sync metadata
                conn.execute(
                    """
                    INSERT OR REPLACE INTO sync_meta (file_path, file_type, file_mtime, file_size, last_synced)
                    VALUES (?, 'sales', ?, ?, ?)
                    """,
                    (file_key, mtime, size, datetime.now().isoformat()),
                )
                conn.commit()
                stats["sales_files_synced"] += 1
                stats["sales_rows_inserted"] += len(rows)

        # 2. Sync Budget & Visitor
        active_b_file = _get_active_budget_file()
        active_m_file = _get_active_mapping_file() or MAPPING_FILE
        if active_b_file and active_b_file.exists():
            b_stat = active_b_file.stat()
            b_key = str(active_b_file.resolve())
            b_mtime = int(b_stat.st_mtime_ns)
            b_size = int(b_stat.st_size)

            cursor = conn.execute(
                "SELECT file_mtime, file_size FROM sync_meta WHERE file_path = ?", (b_key,)
            )
            b_meta = cursor.fetchone()

            if force or not b_meta or b_meta["file_mtime"] != b_mtime or b_meta["file_size"] != b_size:
                print(f"🔄 Sinkronisasi file Budget & Visitor ({active_b_file.name}) ke DB ...")
                targets, visitors = parse_budget_and_visitor(active_b_file, active_m_file)

                conn.execute("BEGIN TRANSACTION;")
                conn.execute("DELETE FROM budget_daily;")
                conn.executemany(
                    """
                    INSERT INTO budget_daily (date, month, outlet, area, target)
                    VALUES (:date, :month, :outlet, :area, :target)
                    """,
                    targets,
                )

                conn.execute("DELETE FROM visitor_actual;")
                conn.executemany(
                    """
                    INSERT INTO visitor_actual (date, month, unit, area, visitors)
                    VALUES (:date, :month, :unit, :area, :visitors)
                    """,
                    visitors,
                )

                conn.execute(
                    """
                    INSERT OR REPLACE INTO sync_meta (file_path, file_type, file_mtime, file_size, last_synced)
                    VALUES (?, 'budget', ?, ?, ?)
                    """,
                    (b_key, b_mtime, b_size, datetime.now().isoformat()),
                )
                conn.commit()
                stats["budget_synced"] = True
                stats["visitors_count"] = len(visitors)
    finally:
        conn.close()

    return stats


if __name__ == "__main__":
    res = sync_database(force=True)
    print("Hasil ETL Database Sync:", res)
